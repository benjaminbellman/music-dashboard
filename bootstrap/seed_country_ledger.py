#!/usr/bin/env python3
"""One-off: read the Artist → Country ledger from the legacy xlsm's Itunes_Info sheet,
normalize codes to ISO 3166-1 alpha-2, and write data/artist_country_seed.csv.

The user's workbook mixes alpha-2 (FR, JP, BE), alpha-3 (USA, SWE, NOR), and
user-specific abbreviations (UK, AUS). This script canonicalizes everything.

Usage:
    python bootstrap/seed_country_ledger.py [path-to-xlsm]
Default: ~/Downloads/Itunes_Dashboard_26.xlsm
"""

from __future__ import annotations

import csv
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_XLSM = Path.home() / "Downloads" / "Itunes_Dashboard_26.xlsm"
OUT_CSV = Path(__file__).resolve().parent.parent / "data" / "artist_country_seed.csv"

# User-specific code → ISO 3166-1 alpha-2.
# Everything not in this map (and not already 2 letters) is flagged as unknown.
CODE_MAP: dict[str, str] = {
    # already-2-letter alpha-2 are passed through untouched if length 2
    # alpha-3 → alpha-2
    "USA": "US",
    "SWE": "SE",
    "NOR": "NO",
    "ITA": "IT",
    "ESP": "ES",
    "BRA": "BR",
    "RUS": "RU",
    "MEX": "MX",
    "COL": "CO",
    "NGA": "NG",
    "UKR": "UA",
    "GRC": "GR",
    "KAZ": "KZ",
    "FIN": "FI",
    "BLR": "BY",
    "ARG": "AR",
    "AUT": "AT",
    "JAM": "JM",
    "LBR": "LR",
    "TUR": "TR",
    "VEN": "VE",
    "PAN": "PA",
    "EGY": "EG",
    "MAR": "MA",
    # user-specific abbreviations
    "UK": "GB",
    "AUS": "AU",
    "MOL": "MD",  # Moldova (user abbrev)
    "SLO": "SI",  # assume Slovenia; SVK-Slovakia is also possible — flag if wrong
}


def normalize(raw: str) -> str | None:
    code = raw.strip().upper()
    if not code or code in ("#N/A", "ND"):
        return None
    # Check the override map FIRST so e.g. "UK" → "GB" instead of
    # falling through as a valid-looking 2-letter code.
    if code in CODE_MAP:
        return CODE_MAP[code]
    if len(code) == 2:
        return code
    return None


def main() -> None:
    xlsm_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSM
    if not xlsm_path.exists():
        raise SystemExit(f"xlsm not found: {xlsm_path}")

    print(f"reading {xlsm_path}")
    wb = load_workbook(str(xlsm_path), data_only=True, read_only=True)
    ws = wb["Itunes_Info"]

    seen: dict[str, str] = {}
    raw_counts: Counter[str] = Counter()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 5:
            continue
        artist = row[4]  # column E
        country = row[3]  # column D
        if not artist:
            continue
        a = str(artist).strip()
        if not a or a in seen:
            continue
        raw_counts[(str(country).strip() if country else "")] += 1
        iso = normalize(str(country) if country else "")
        if iso:
            seen[a] = iso

    # Anything we couldn't normalize — report so the user can audit
    unresolved = {
        code: n for code, n in raw_counts.items() if code and normalize(code) is None
    }

    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["artist", "country"])
        for artist in sorted(seen):
            w.writerow([artist, seen[artist]])

    print(f"wrote {len(seen):,} artist→country rows to {OUT_CSV}")
    if unresolved:
        print("\nunresolved source codes (these artists were skipped):")
        for code, n in sorted(unresolved.items(), key=lambda x: -x[1]):
            print(f"  {code!r:<10s} {n:>4d} artists")


if __name__ == "__main__":
    main()
